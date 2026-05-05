#!/usr/bin/env python3
"""Import YGE Excel master data into the API.

Usage:
    python3 scripts/import-from-excel.py <path-to-xlsx> [--api-base URL]

Default API: https://yge-api.onrender.com

What it imports:
  - Cost_Codes        -> POST /api/cost-codes
  - Equipment_Rates   -> POST /api/equipment-rates  (kind=OWNED)
  - Equipment_Rental  -> POST /api/equipment-rates  (kind=RENTAL)
  - Est_*             -> POST /api/imported-estimates  (one per sheet)
"""

import argparse
import json
import sys
from typing import Any, Optional
import urllib.request
import urllib.error

import openpyxl


def to_cents(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(round(float(v) * 100))
    except (TypeError, ValueError):
        return 0


def to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def post_json(url: str, body: dict) -> dict:
    data = json.dumps(body).encode("utf-8")
    req = urllib.request.Request(
        url, data=data, headers={"Content-Type": "application/json"}, method="POST"
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        msg = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"POST {url} -> HTTP {e.code}: {msg[:300]}") from e


def import_cost_codes(wb: openpyxl.Workbook, api_base: str) -> int:
    ws = wb["Cost_Codes"]
    count = 0
    SOURCE_MAP = {
        "Labor_Rates": "Labor_Rates",
        "Equipment_Rates": "Equipment_Rates",
        "Equipment_Rental": "Equipment_Rental",
        "Materials": "Materials",
        "Subcontractors": "Subcontractors",
    }
    # Header at row 2; data starts row 3.
    for r in range(3, ws.max_row + 1):
        code = ws.cell(r, 1).value
        if not code:
            continue
        category = ws.cell(r, 2).value
        description = ws.cell(r, 3).value
        rate_source = ws.cell(r, 4).value
        body = {
            "code": str(code).strip(),
            "rateSource": SOURCE_MAP.get(str(rate_source).strip() if rate_source else "", "Other"),
        }
        if category: body["category"] = str(category).strip()
        if description: body["description"] = str(description).strip()
        post_json(f"{api_base}/api/cost-codes", body)
        count += 1
    print(f"  cost codes: {count}")
    return count


def import_equipment_rates(wb: openpyxl.Workbook, api_base: str) -> int:
    count = 0

    # OWNED — header at row 4, data row 5+.
    ws = wb["Equipment_Rates"]
    for r in range(5, ws.max_row + 1):
        cost_code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if not cost_code or not name:
            continue
        body = {
            "costCode": str(cost_code).strip(),
            "name": str(name).strip(),
            "kind": "OWNED",
            "bareRateCents": to_cents(ws.cell(r, 3).value),
            "gallonsPerHour": to_float(ws.cell(r, 4).value),
            "fuelCentsPerHour": to_cents(ws.cell(r, 5).value),
            "totalCentsPerHour": to_cents(ws.cell(r, 6).value),
            "unit": str(ws.cell(r, 7).value or "hr"),
        }
        notes = ws.cell(r, 8).value
        if notes: body["notes"] = str(notes).strip()
        post_json(f"{api_base}/api/equipment-rates", body)
        count += 1

    # RENTAL — header at row 3, data row 4+.
    ws = wb["Equipment_Rental"]
    SOURCE_MAP = {"Conf": "Confirmed", "Est": "Estimated"}
    for r in range(4, ws.max_row + 1):
        cost_code = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        if not cost_code or not name:
            continue
        category = ws.cell(r, 3).value
        body = {
            "costCode": str(cost_code).strip(),
            "name": str(name).strip(),
            "kind": "RENTAL",
            "dailyCents": to_cents(ws.cell(r, 4).value),
            "weeklyCents": to_cents(ws.cell(r, 5).value),
            "monthlyCents": to_cents(ws.cell(r, 6).value),
            "source": SOURCE_MAP.get(str(ws.cell(r, 7).value).strip() if ws.cell(r, 7).value else "", "Other"),
        }
        if category: body["category"] = str(category).strip()
        notes = ws.cell(r, 8).value
        if notes: body["notes"] = str(notes).strip()
        post_json(f"{api_base}/api/equipment-rates", body)
        count += 1

    print(f"  equipment rates: {count}")
    return count


def import_estimate(wb: openpyxl.Workbook, sheet_name: str, api_base: str) -> Optional[str]:
    ws = wb[sheet_name]
    if ws.max_row < 10:
        return None

    # Project header (rows 1-7).
    job_number = str(ws.cell(3, 5).value or sheet_name.replace("Est_", "")).strip()
    project_name = str(ws.cell(3, 6).value or sheet_name).strip()
    rate_type_raw = str(ws.cell(3, 8).value or "PW").strip()
    rate_type = "Private" if rate_type_raw.lower().startswith("priv") else "PW"
    opp_percent = to_float(ws.cell(3, 10).value) or 0.20

    direct_cents = to_cents(ws.cell(6, 5).value)
    opp_cents = to_cents(ws.cell(6, 9).value)
    bid_cents = to_cents(ws.cell(6, 12).value)

    # Find the project's job number → look up the client from the Jobs sheet.
    client = ""
    jobs = wb["Jobs"]
    for r in range(3, jobs.max_row + 1):
        if str(jobs.cell(r, 1).value or "").strip() == job_number:
            client = str(jobs.cell(r, 3).value or "").strip()
            break

    CATEGORY_MAP = {
        "Labor": "LABOR",
        "Equipment (Owned)": "EQUIPMENT_OWNED",
        "Equipment (Rental)": "EQUIPMENT_RENTAL",
        "Material": "MATERIAL",
        "Subcontract": "SUBCONTRACT",
        "Other": "OTHER",
    }

    lines = []
    current_section = ""
    # Data starts at row 10.
    for r in range(10, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col4 = ws.cell(r, 4).value  # Category column

        # Section header row: text in col 1, no Category set on row.
        if col1 is not None and not isinstance(col1, (int, float)) and not col4:
            text = str(col1).strip()
            if text and text.upper() == text and len(text) > 5:
                current_section = text
                continue

        # Line item row: col 1 is an integer item #.
        if not isinstance(col1, (int, float)):
            continue

        category = str(col4 or "").strip()
        if category not in CATEGORY_MAP:
            continue

        line = {
            "itemNumber": int(col1),
            "category": CATEGORY_MAP[category],
            "description": str(ws.cell(r, 6).value or "").strip(),
            "quantity": to_float(ws.cell(r, 7).value),
            "otMultiplier": to_float(ws.cell(r, 9).value) or 1,
            "unitCostCents": to_cents(ws.cell(r, 10).value),
            "totalCostCents": to_cents(ws.cell(r, 11).value),
            "oppMarkupCents": to_cents(ws.cell(r, 12).value),
            "bidPriceCents": to_cents(ws.cell(r, 13).value),
        }
        if current_section: line["sectionName"] = current_section
        cost_code = ws.cell(r, 5).value
        if cost_code: line["costCode"] = str(cost_code).strip()
        unit = ws.cell(r, 8).value
        if unit: line["unit"] = str(unit).strip()
        notes = ws.cell(r, 14).value
        if notes: line["notes"] = str(notes).strip()

        lines.append(line)

    body = {
        "jobNumber": job_number,
        "projectName": project_name,
        "rateType": rate_type,
        "oppPercent": opp_percent,
        "directCostCents": direct_cents,
        "oppMarkupCents": opp_cents,
        "bidPriceCents": bid_cents,
        "lines": lines,
    }
    if client: body["client"] = client

    result = post_json(f"{api_base}/api/imported-estimates", body)
    eid = result.get("importedEstimate", {}).get("id")
    print(f"  estimate {sheet_name}: {len(lines)} lines, id={eid}")
    return eid


def import_estimates(wb: openpyxl.Workbook, api_base: str) -> int:
    count = 0
    for sheet in wb.sheetnames:
        if sheet.startswith("Est_"):
            try:
                if import_estimate(wb, sheet, api_base):
                    count += 1
            except Exception as e:  # pragma: no cover
                print(f"    skipped {sheet}: {e}")
    return count


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path")
    parser.add_argument("--api-base", default="https://yge-api.onrender.com")
    args = parser.parse_args()

    print(f"Loading {args.xlsx_path}")
    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)
    print(f"API base: {args.api_base}")

    print("Importing cost codes…")
    import_cost_codes(wb, args.api_base)
    print("Importing equipment rates…")
    import_equipment_rates(wb, args.api_base)
    print("Importing estimates…")
    import_estimates(wb, args.api_base)

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
